from fastapi import FastAPI, APIRouter, File, UploadFile, HTTPException
from fastapi.responses import StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
from dotenv import load_dotenv
from motor.motor_asyncio import AsyncIOMotorClient
import os
import httpx
import logging
import io
import base64
import uuid
import json
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Optional, Dict, Any
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from PIL import Image as PILImage
import pillow_heif

# Register HEIF opener with Pillow to support iPhone images (HEIC format) natively
pillow_heif.register_heif_opener()

# Global In-Memory Database Fallback if MongoDB is offline
IN_MEMORY_DB = []

# High-efficiency Image Compression to prevent 400 Bad Request Payload Size Limits
def compress_image(image_bytes: bytes, max_size: int = 1024) -> bytes:
    try:
        # Load image
        img = PILImage.open(io.BytesIO(image_bytes))
        
        # Convert to RGB if in RGBA mode (PNGs) to support JPEG saving
        if img.mode in ('RGBA', 'LA') or (img.mode == 'P' and 'transparency' in img.info):
            background = PILImage.new("RGB", img.size, (255, 255, 255))
            background.paste(img, mask=img.split()[3] if img.mode == 'RGBA' else None)
            img = background
        elif img.mode != 'RGB':
            img = img.convert('RGB')
            
        # Downscale if larger than max_size (keeps exceptional OCR legibility while reducing size 95%)
        img.thumbnail((max_size, max_size), PILImage.Resampling.LANCZOS)
        
        # Save to buffer with high compression
        buffer = io.BytesIO()
        img.save(buffer, format="JPEG", quality=70, optimize=True)
        compressed = buffer.getvalue()
        logging.info(f"⚡ Image compressed from {len(image_bytes)} bytes to {len(compressed)} bytes")
        return compressed
    except Exception as e:
        logging.warning(f"⚠️ Failed to compress image, using original: {e}")
        return image_bytes

# Try EmergentIntegrations
try:
    from emergentintegrations.llm.chat import LlmChat, UserMessage, ImageContent
    EMERGENT_AVAILABLE = True
except ImportError:
    EMERGENT_AVAILABLE = False
    print("⚠️  EmergentIntegrations not found. Install with:")
    print("pip install emergentintegrations --extra-index-url https://d33sy5i8bnduwe.cloudfront.net/simple/")

# Load environment variables
ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ.get('MONGO_URL', 'mongodb://localhost:27017')
db_name = os.environ.get('DB_NAME', 'handwritten_tables')
client = AsyncIOMotorClient(mongo_url)
db = client[db_name]

# Create FastAPI app
app = FastAPI(title="Handwritten Table Converter", version="1.0.0")
api_router = APIRouter(prefix="/api")

# ✅ FIXED CORS Middleware (allow all during dev)
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],  # for dev, allow all origins
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

# -------------------------
# Models
# -------------------------
class TableData(BaseModel):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    filename: str
    extracted_data: List[List[str]]
    created_at: datetime = Field(default_factory=datetime.utcnow)

class ProcessingResult(BaseModel):
    success: bool
    message: str
    table_data: Optional[List[List[str]]] = None
    processing_id: Optional[str] = None

# -------------------------
# Utility Functions
# -------------------------
def image_to_base64(image_bytes: bytes) -> str:
    return base64.b64encode(image_bytes).decode('utf-8')

async def extract_table_from_image(image_bytes: bytes, filename: str) -> Dict[str, Any]:
    try:
        api_key = os.environ.get('EMERGENT_LLM_KEY')
        image_base64 = image_to_base64(image_bytes)

        # 1. Fallback 1: Direct Vision API (supports both OpenRouter Free & standard OpenAI keys)
        if api_key:
            # Since compress_image always outputs a high-efficiency JPEG, MIME type is always image/jpeg
            mime_type = "image/jpeg"

            # Determine endpoint, headers, and model fallback list based on API key prefix
            if api_key.startswith('sk-or-') or 'openrouter' in api_key.lower():
                logging.info("🧠 Using OpenRouter Vision API with free-tier model fallbacks")
                endpoint = "https://openrouter.ai/api/v1/chat/completions"
                headers = {
                    "Authorization": f"Bearer {api_key}",
                    "Content-Type": "application/json",
                    "HTTP-Referer": "https://github.com/verupakshasharma9/Handwritten-Excel-Converter",
                    "X-Title": "Handwritten Table Converter"
                }
                models_to_try = [
                    "google/gemini-flash-1.5:free",                   # 100% Free
                    "qwen/qwen-2-vl-7b-instruct:free",               # 100% Free
                    "meta-llama/llama-3.2-11b-vision-instruct:free"   # 100% Free
                ]
            else:
                logging.info("🧠 Using standard OpenAI Vision API with gpt-4o-mini vision fallback")
                endpoint = "https://api.openai.com/v1/chat/completions"
                headers = {
                    "Authorization": f"Bearer {api_key}",
                    "Content-Type": "application/json"
                }
                models_to_try = [
                    "gpt-4o-mini",                                    # OpenAI trial-friendly, ultra-cheap vision model
                    "gpt-4o"
                ]

            last_error = None
            for model_name in models_to_try:
                try:
                    logging.info(f"🚀 Attempting table extraction using model: {model_name}")
                    payload = {
                        "model": model_name,
                        "messages": [
                            {
                                "role": "user",
                                "content": [
                                    {
                                        "type": "text",
                                        "text": "You are an expert at analyzing handwritten tables and extracting structured data.\n\nAnalyze this handwritten table image and extract all data into structured format.\n\nReturn ONLY a valid JSON array of arrays representing the table, containing the header row first and data rows. Do NOT wrap in markdown code fences, do not write ```json, just return pure JSON text."
                                    },
                                    {
                                        "type": "image_url",
                                        "image_url": {
                                            "url": f"data:{mime_type};base64,{image_base64}"
                                        }
                                    }
                                ]
                            }
                        ]
                    }
                    async with httpx.AsyncClient(timeout=60.0) as client:
                        response = await client.post(endpoint, headers=headers, json=payload)
                        response.raise_for_status()
                        result = response.json()
                        content = result["choices"][0]["message"]["content"].strip()
                        
                        # Clean content if it contains markdown code fences
                        if content.startswith("```"):
                            content = content.strip("`json").strip("`").strip()
                        
                        table_data = json.loads(content)
                        if not isinstance(table_data, list) or not table_data:
                            raise ValueError("Parsed data is not a list")

                        logging.info(f"✅ Successful table extraction with model: {model_name}")
                        return {
                            "success": True,
                            "table_data": table_data,
                            "message": f"Table extracted successfully via OpenRouter ({model_name})"
                        }
                except httpx.HTTPStatusError as http_err:
                    err_body = http_err.response.text
                    logging.error(f"❌ OpenRouter HTTP Error {http_err.response.status_code} for {model_name}: {err_body}")
                    last_error = http_err
                    continue
                except Exception as e:
                    logging.warning(f"⚠️ Model {model_name} failed: {e}")
                    last_error = e
                    continue

            # If all models in the fallback loop fail, log it and return the exact API error details gracefully
            err_msg = f"All models failed. Last error: {str(last_error)}"
            if isinstance(last_error, httpx.HTTPStatusError):
                err_msg = f"OpenRouter Error: {last_error.response.text}"
            
            logging.error(f"❌ {err_msg}")
            return {
                "success": False,
                "message": err_msg,
                "table_data": None
            }

        # 2. Standard Emergent Integrations library
        if EMERGENT_AVAILABLE and api_key:
            logging.info("🧠 Using EmergentIntegrations library for table extraction")
            chat = LlmChat(
                api_key=api_key,
                session_id=f"table_extraction_{uuid.uuid4()}",
                system_message="You are an expert at analyzing handwritten tables and extracting structured data."
            ).with_model("openai", "gpt-4o")

            image_content = ImageContent(image_base64=image_base64)
            prompt = """Analyze this handwritten table image and extract all data into structured format.

Return ONLY valid JSON array like:
[
  ["Header1", "Header2"],
  ["Row1Col1", "Row1Col2"]
]"""

            user_message = UserMessage(text=prompt, file_contents=[image_content])
            response = await chat.send_message(user_message)
            response_text = response.strip()
            if response_text.startswith('```'):
                response_text = response_text.strip('`json').strip('`')
            table_data = json.loads(response_text.strip())

            return {
                "success": True,
                "table_data": table_data,
                "message": "Table extracted successfully"
            }

        # 3. Local Fallback / Mock Data if no key is set or library unavailable
        logging.info("⚠️ No API key or library found. Using local mock OCR data.")
        return {
            "success": True,
            "table_data": [
                ["Name", "Age", "City"],
                ["John", "25", "NYC"],
                ["Alice", "30", "LA"]
            ],
            "message": "Table extracted successfully (Mock data fallback)"
        }

    except Exception as e:
        logging.error(f"OCR processing error: {e}")
        return {"success": False, "message": f"Error: {str(e)}", "table_data": None}

def create_excel_file(table_data: List[List[str]], filename: str) -> io.BytesIO:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Extracted Table"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    border = Border(left=Side(style='thin'), right=Side(style='thin'),
                    top=Side(style='thin'), bottom=Side(style='thin'))
    center_alignment = Alignment(horizontal='center', vertical='center')

    for row_idx, row_data in enumerate(table_data, 1):
        for col_idx, cell_value in enumerate(row_data, 1):
            cell = worksheet.cell(row=row_idx, column=col_idx, value=cell_value)
            cell.border = border
            cell.alignment = center_alignment
            if row_idx == 1:
                cell.font = header_font
                cell.fill = header_fill

    for column in worksheet.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        worksheet.column_dimensions[column_letter].width = min(max_length + 2, 50)

    excel_buffer = io.BytesIO()
    workbook.save(excel_buffer)
    excel_buffer.seek(0)
    return excel_buffer

# -------------------------
# API Routes
# -------------------------
@api_router.get("/")
async def root():
    return {"message": "Handwritten Table Converter API", "status": "running"}

@api_router.post("/upload-image", response_model=ProcessingResult)
async def upload_image(file: UploadFile = File(...)):
    if not file.content_type.startswith('image/'):
        raise HTTPException(status_code=400, detail="Only image files are allowed")

    image_bytes = await file.read()
    compressed_bytes = compress_image(image_bytes)
    result = await extract_table_from_image(compressed_bytes, file.filename)

    if result["success"]:
        table_record = TableData(filename=file.filename, extracted_data=result["table_data"])
        try:
            await db.table_extractions.insert_one(table_record.dict())
        except Exception as e:
            logging.warning(f"⚠️ MongoDB write failed, using local in-memory DB: {e}")
            IN_MEMORY_DB.append(table_record.dict())

        return ProcessingResult(success=True, message=result["message"],
                                table_data=result["table_data"], processing_id=table_record.id)
    else:
        return ProcessingResult(success=False, message=result["message"])

@api_router.post("/generate-excel/{processing_id}")
async def generate_excel(processing_id: str):
    record = None
    try:
        record = await db.table_extractions.find_one({"id": processing_id})
    except Exception as e:
        logging.warning(f"⚠️ MongoDB read failed, falling back to in-memory DB: {e}")

    if not record:
        # Search in memory list
        record = next((r for r in IN_MEMORY_DB if r["id"] == processing_id), None)

    if not record:
        raise HTTPException(status_code=404, detail="Processing record not found")

    excel_buffer = create_excel_file(record["extracted_data"], record["filename"])
    excel_filename = f"{record['filename'].rsplit('.',1)[0]}_extracted.xlsx"

    return StreamingResponse(
        io.BytesIO(excel_buffer.read()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={excel_filename}"}
    )

@api_router.get("/extractions", response_model=List[TableData])
async def get_extractions():
    records = []
    try:
        records = await db.table_extractions.find().sort("created_at", -1).to_list(50)
    except Exception as e:
        logging.warning(f"⚠️ MongoDB list failed, reading from in-memory DB: {e}")
        records = sorted(IN_MEMORY_DB, key=lambda x: x.get("created_at", datetime.utcnow()), reverse=True)[:50]

    return [TableData(**record) for record in records]

@app.get("/")
async def root_health():
    return {"message": "Handwritten Table Converter Backend is active and running", "status": "healthy"}

# Include router
app.include_router(api_router)

# Logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')

@app.on_event("startup")
async def startup_event():
    logging.info("🚀 API Started")
    logging.info(f"📊 Database: {mongo_url}/{db_name}")
    logging.info("🧠 AI Integration")

@app.on_event("shutdown")
async def shutdown_event():
    try:
        client.close()
    except Exception:
        pass
    logging.info("📴 Application shutdown complete")

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="0.0.0.0", port=8000)
